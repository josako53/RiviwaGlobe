"""
core/exporters.py
═══════════════════════════════════════════════════════════════════════════════
PDF and Excel export helpers for all report endpoints.

Usage in any report endpoint:
    from core.exporters import export_response

    @router.get("/grievances")
    async def grievance_performance(..., format: str = Query(default="json")):
        data = { ... }          # your existing dict
        return export_response(data, format=format, filename="grievances")

export_response() returns:
  format="json" → the dict unchanged (FastAPI serialises it)
  format="pdf"  → StreamingResponse with application/pdf
  format="xlsx" → StreamingResponse with application/vnd.openxmlformats...

PDF layout:
  - Cover: title + date range + applied filters
  - One section per top-level key
  - Tables for list fields (by_status, by_priority, by_channel, etc.)
  - Metric cards for scalar fields (total, resolution_rate, etc.)

Excel layout:
  - Sheet "Summary"  — all scalar metrics
  - One sheet per list field (by_status, by_priority, by_channel, etc.)
  - Sheet "Log"      — if "items" key present (full log rows)
═══════════════════════════════════════════════════════════════════════════════
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Dict, List, Optional, Union

from fastapi.responses import StreamingResponse


# ─────────────────────────────────────────────────────────────────────────────
# Public entry point
# ─────────────────────────────────────────────────────────────────────────────

def export_response(
    data:     Dict[str, Any],
    format:   str,
    filename: str,
    title:    Optional[str] = None,
) -> Any:
    """
    Returns the data as JSON (passthrough), PDF StreamingResponse,
    or Excel StreamingResponse depending on `format`.

    format values: "json" | "pdf" | "xlsx" | "csv"
    filename: used as the download filename (without extension)
    title: optional human-readable report title for PDF header
    """
    fmt = format.lower().strip()
    if fmt in ("pdf",):
        buf = _build_pdf(data, title or filename.replace("-", " ").title())
        return StreamingResponse(
            buf,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}.pdf"'},
        )
    if fmt in ("xlsx", "excel"):
        buf = _build_xlsx(data, title or filename.replace("-", " ").title())
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
        )
    if fmt == "csv":
        # CSV only makes sense for log-style reports (data["items"])
        buf = _build_csv(data)
        return StreamingResponse(
            buf,
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
        )
    # Default: return dict (FastAPI JSONResponse)
    return data


# ─────────────────────────────────────────────────────────────────────────────
# PDF builder
# ─────────────────────────────────────────────────────────────────────────────

def _build_pdf(data: Dict[str, Any], title: str) -> io.BytesIO:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        HRFlowable, PageBreak, Paragraph, SimpleDocTemplate,
        Spacer, Table, TableStyle,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=20*mm, rightMargin=20*mm,
        topMargin=20*mm, bottomMargin=20*mm,
    )

    styles = getSampleStyleSheet()
    H1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=18, spaceAfter=6)
    H2 = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=13, spaceAfter=4, spaceBefore=10)
    BODY  = ParagraphStyle("BODY",  parent=styles["Normal"], fontSize=9, spaceAfter=3)
    LABEL = ParagraphStyle("LABEL", parent=styles["Normal"], fontSize=8,
                           textColor=colors.HexColor("#666666"), spaceAfter=2)

    # Brand colours
    BRAND_DARK  = colors.HexColor("#0C447C")
    BRAND_MID   = colors.HexColor("#E6F1FB")
    ROW_ALT     = colors.HexColor("#F7FAFE")
    HEADER_TEXT = colors.white

    story = []

    # ── Cover header ──────────────────────────────────────────────────────────
    story.append(Paragraph(title, H1))

    # Date range + metadata
    dr = data.get("date_range", {})
    if dr:
        from_str = dr.get("from", "")[:10]
        to_str   = dr.get("to",   "")[:10]
        story.append(Paragraph(f"Period: {from_str} to {to_str}", LABEL))

    if pid := data.get("project_id"):
        story.append(Paragraph(f"Project: {pid}", LABEL))

    filters = data.get("applied_filters", {})
    if filters:
        f_str = "  ·  ".join(f"{k}: {v}" for k, v in filters.items())
        story.append(Paragraph(f"Filters: {f_str}", LABEL))

    story.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", LABEL))
    story.append(HRFlowable(width="100%", thickness=1, color=BRAND_DARK, spaceAfter=8))

    # ── Helper: render a list-of-dicts as a table ──────────────────────────────
    def _table(rows: List[Dict[str, Any]], heading: str) -> None:
        if not rows:
            return
        story.append(Paragraph(heading.replace("_", " ").title(), H2))
        keys = list(rows[0].keys())
        header_row = [k.replace("_", " ").title() for k in keys]
        table_data = [header_row]
        for row in rows:
            table_data.append([str(row.get(k, "")) for k in keys])
        col_w = (A4[0] - 40*mm) / len(keys)
        t = Table(table_data, colWidths=[col_w] * len(keys), repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), BRAND_DARK),
            ("TEXTCOLOR",  (0,0), (-1,0), HEADER_TEXT),
            ("FONTSIZE",   (0,0), (-1,0), 8),
            ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",   (0,1), (-1,-1), 8),
            ("FONTNAME",   (0,1), (-1,-1), "Helvetica"),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, ROW_ALT]),
            ("GRID",       (0,0), (-1,-1), 0.3, colors.HexColor("#CCCCCC")),
            ("ALIGN",      (0,0), (-1,-1), "LEFT"),
            ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
            ("TOPPADDING", (0,0), (-1,-1), 3),
            ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ]))
        story.append(t)
        story.append(Spacer(1, 6))

    # ── Helper: render scalar metrics in a 2-col table ─────────────────────────
    def _metrics(d: Dict[str, Any], heading: str) -> None:
        scalars = {k: v for k, v in d.items()
                   if isinstance(v, (int, float, str, type(None)))
                   and k not in ("project_id", "date_range", "applied_filters")}
        if not scalars:
            return
        story.append(Paragraph(heading, H2))
        rows = [[k.replace("_", " ").title(), str(v) if v is not None else "—"]
                for k, v in scalars.items()]
        col_w = (A4[0] - 40*mm) / 2
        t = Table(rows, colWidths=[col_w, col_w])
        t.setStyle(TableStyle([
            ("FONTSIZE",   (0,0), (-1,-1), 8),
            ("FONTNAME",   (0,0), (0,-1), "Helvetica-Bold"),
            ("FONTNAME",   (1,0), (1,-1), "Helvetica"),
            ("ROWBACKGROUNDS", (0,0), (-1,-1), [colors.white, ROW_ALT]),
            ("GRID",       (0,0), (-1,-1), 0.3, colors.HexColor("#CCCCCC")),
            ("ALIGN",      (1,0), (1,-1), "RIGHT"),
            ("TOPPADDING", (0,0), (-1,-1), 3),
            ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ]))
        story.append(t)
        story.append(Spacer(1, 6))

    # ── Render top-level fields ────────────────────────────────────────────────
    # Metrics section first
    _metrics(data, "Key metrics")

    # Then every list field becomes its own table
    skip_keys = {"project_id", "date_range", "applied_filters", "recent_breaches"}
    for key, val in data.items():
        if key in skip_keys:
            continue
        if isinstance(val, list) and val and isinstance(val[0], dict):
            _table(val, key)
        elif isinstance(val, dict):
            # nested dict — render as its own metrics section
            inner_lists = {k: v for k, v in val.items() if isinstance(v, list) and v and isinstance(v[0], dict)}
            inner_scalars = {k: v for k, v in val.items() if not isinstance(v, list)}
            if inner_scalars:
                _metrics(inner_scalars, key.replace("_", " ").title())
            for ik, iv in inner_lists.items():
                _table(iv, f"{key} — {ik}")

    # Breaches at end (can be long)
    if breaches := data.get("recent_breaches"):
        _table(breaches, "SLA breach items")

    doc.build(story)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────────────────
# Excel builder
# ─────────────────────────────────────────────────────────────────────────────

def _build_xlsx(data: Dict[str, Any], title: str) -> io.BytesIO:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # Colour constants
    BRAND_DARK = "0C447C"
    BRAND_LIGHT = "E6F1FB"
    ALT_ROW    = "F7FAFE"
    WHITE      = "FFFFFF"

    def _header_style(cell):
        cell.font      = Font(bold=True, color=WHITE, size=10)
        cell.fill      = PatternFill("solid", fgColor=BRAND_DARK)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def _subheader_style(cell):
        cell.font      = Font(bold=True, color="0C447C", size=10)
        cell.fill      = PatternFill("solid", fgColor=BRAND_LIGHT)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    def _write_list_sheet(name: str, rows: List[Dict[str, Any]]) -> None:
        if not rows:
            return
        safe_name = name[:31].replace("/", "-").replace("\\", "-").replace("*","").replace("?","").replace("[","").replace("]","")
        ws = wb.create_sheet(title=safe_name)
        keys = list(rows[0].keys())
        # Header
        for ci, k in enumerate(keys, 1):
            cell = ws.cell(row=1, column=ci, value=k.replace("_", " ").title())
            _header_style(cell)
            cell.border = border
        # Rows
        for ri, row in enumerate(rows, 2):
            fill_color = WHITE if ri % 2 == 0 else ALT_ROW
            for ci, k in enumerate(keys, 1):
                cell = ws.cell(row=ri, column=ci, value=row.get(k))
                cell.fill      = PatternFill("solid", fgColor=fill_color)
                cell.font      = Font(size=9)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border    = border
        ws.freeze_panes = "A2"
        _auto_width(ws)

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    # Title
    ws_sum.merge_cells("A1:D1")
    tc = ws_sum["A1"]
    tc.value     = title
    tc.font      = Font(bold=True, size=14, color=BRAND_DARK)
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws_sum.row_dimensions[1].height = 24

    # Metadata
    row = 2
    dr = data.get("date_range", {})
    if dr:
        ws_sum.cell(row=row, column=1, value="Period").font = Font(bold=True, size=9, color="666666")
        ws_sum.cell(row=row, column=2, value=f"{dr.get('from','')[:10]} to {dr.get('to','')[:10]}").font = Font(size=9)
        row += 1
    if pid := data.get("project_id"):
        ws_sum.cell(row=row, column=1, value="Project").font = Font(bold=True, size=9, color="666666")
        ws_sum.cell(row=row, column=2, value=str(pid)).font = Font(size=9)
        row += 1
    ws_sum.cell(row=row, column=1, value="Generated").font = Font(bold=True, size=9, color="666666")
    ws_sum.cell(row=row, column=2, value=datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")).font = Font(size=9)
    row += 2

    # All scalar metrics
    skip = {"project_id", "date_range", "applied_filters"}
    for key, val in data.items():
        if key in skip:
            continue
        if isinstance(val, (int, float, str)) and val is not None:
            lbl = ws_sum.cell(row=row, column=1, value=key.replace("_", " ").title())
            lbl.font      = Font(bold=True, size=9)
            lbl.alignment = Alignment(horizontal="left")
            vc = ws_sum.cell(row=row, column=2, value=val)
            vc.font      = Font(size=9)
            vc.alignment = Alignment(horizontal="right")
            row += 1
        elif isinstance(val, dict):
            # nested metrics
            row += 1
            hc = ws_sum.cell(row=row, column=1, value=key.replace("_", " ").title())
            _subheader_style(hc)
            row += 1
            for k2, v2 in val.items():
                if isinstance(v2, (int, float, str, type(None))):
                    ws_sum.cell(row=row, column=1, value=f"  {k2.replace('_',' ').title()}").font = Font(size=9, italic=True)
                    ws_sum.cell(row=row, column=2, value=v2).font = Font(size=9)
                    row += 1

    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 20

    # ── One sheet per list field ───────────────────────────────────────────────
    list_keys = [
        "by_status", "by_priority", "by_channel", "by_submission_method",
        "by_level", "by_lga", "by_type", "by_period",
        "escalation_breakdown", "sla_by_priority", "recent_breaches",
        "categories", "items",
    ]
    for key in list_keys:
        val = data.get(key)
        if isinstance(val, list) and val and isinstance(val[0], dict):
            _write_list_sheet(key.replace("_", " ").title(), val)

    # Nested list fields (e.g. sla_compliance contains sub-lists)
    for key, val in data.items():
        if isinstance(val, dict):
            for k2, v2 in val.items():
                if isinstance(v2, list) and v2 and isinstance(v2[0], dict):
                    _write_list_sheet(f"{key[:15]}-{k2[:14]}", v2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────────────────
# CSV builder (log reports only)
# ─────────────────────────────────────────────────────────────────────────────

def _build_csv(data: Dict[str, Any]) -> io.BytesIO:
    import csv

    items = data.get("items", [])
    if not items:
        # Fall back to scalar metrics
        items = [{k: v for k, v in data.items()
                  if isinstance(v, (int, float, str, type(None)))}]

    buf = io.BytesIO()
    wrapper = io.TextIOWrapper(buf, encoding="utf-8-sig", newline="")  # utf-8-sig for Excel compat
    writer  = csv.DictWriter(wrapper, fieldnames=list(items[0].keys()))
    writer.writeheader()
    writer.writerows(items)
    wrapper.flush()
    wrapper.detach()
    buf.seek(0)
    return buf
